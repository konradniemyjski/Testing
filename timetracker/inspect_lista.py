from openpyxl import load_workbook

wb = load_workbook('/home/konrad/Dokumenty/Testing/lista.xlsx')
ws = wb.active

print(f"Sheet Title: {ws.title}")

print("\n--- Rows 1-3 Values ---")
for row in ws.iter_rows(min_row=1, max_row=3, max_col=40):
    values = [cell.value for cell in row]
    print(values)

print("\n--- Rows 1-3 Styles (Sample) ---")
# Check styles for first few columns and day columns
columns_to_check = [1, 2, 3, 4, 5, 6, 36, 37] # A, B, C, D, E, F... AJ, AK
for row_idx in range(1, 4):
    print(f"Row {row_idx}:")
    for col_idx in columns_to_check:
        cell = ws.cell(row=row_idx, column=col_idx)
        print(f"  Col {col_idx}: Value={cell.value}, FontColor={cell.font.color.rgb if cell.font.color else 'None'}, FillColor={cell.fill.start_color.rgb if cell.fill.start_color else 'None'}")
