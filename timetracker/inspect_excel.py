import openpyxl
import sys

try:
    wb = openpyxl.load_workbook('/home/konrad/Dokumenty/Testing/timetracker/lista.xlsx')
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    
    headers = []
    # Read headers from the first row, up to column AK (which is column 37)
    for col in range(1, 38): 
        cell_value = sheet.cell(row=1, column=col).value
        headers.append(f"{openpyxl.utils.get_column_letter(col)}1: {cell_value}")
    
    for col in range(1, 38): 
        cell_value = sheet.cell(row=2, column=col).value
        headers.append(f"{openpyxl.utils.get_column_letter(col)}2: {cell_value}")
        
    print("\nHeaders:")
    for h in headers:
        print(h)
        
except Exception as e:
    print(f"Error: {e}")
