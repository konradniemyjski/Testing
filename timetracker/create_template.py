from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Lista"

# Set B1
ws["B1"] = "MIESIĄC ROK"
ws["B1"].font = Font(bold=True, size=14)

# Headers Row 1 (Optional, but good for context)
headers = ["Lp", "Nazwisko i Imię", "Stawka/h", ""] + [str(i) for i in range(1, 32)] + ["Razem h", "Do wypłaty"]
# Columns: A=1, B=2, C=3, D=4, E=5...
# E is index 4 in list (0-based)

# Write Row 2 as the "Template" row, but maybe we should put headers in Row 1?
# The code says: "Row 2 is the template row with formulas... Start filling from Row 2."
# And "Start filling from Row 2" implies Row 2 IS the first data row.
# But then it says "We will retain it for the first user... duplicate styles...".
# This implies Row 2 is ALREADY filled or formatted in the template.
# Let's assume Row 1 is Headers. Row 2 is the Formatting Template.

# Let's write headers in Row 1
ws.cell(row=1, column=1, value="Lp")
ws.cell(row=1, column=2, value="Nazwisko i Imię")
ws.cell(row=1, column=3, value="Stawka")
ws.cell(row=1, column=4, value="") # Buffer column
for day in range(1, 32):
    ws.cell(row=1, column=4+day, value=day)
ws.cell(row=1, column=36, value="Razem") # AJ
ws.cell(row=1, column=37, value="Kwota") # AK

# Style Row 2 (Template)
# We don't need to put data, just styles.
medium_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

for col in range(1, 38):
    cell = ws.cell(row=2, column=col)
    cell.border = medium_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if col == 2: # Name
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Add connection formulas for Row 2
ws["AJ2"] = "=SUM(E2:AI2)"
ws["AK2"] = "=AJ2*C2"

# Save
wb.save("backend/templates/lista.xlsx")
print("Template created at backend/templates/lista.xlsx")
