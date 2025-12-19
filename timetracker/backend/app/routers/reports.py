from datetime import datetime
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.orm import joinedload

from .. import auth, models, schemas
from ..routers.worklogs import _build_user_worklog_query
from ..db import get_db

router = APIRouter(prefix="/reports", tags=["reports"])

from enum import Enum

class ReportType(str, Enum):
    PARTICIPANTS = "participants"
    TEAM_WORK = "team_work"
    ACCOMMODATION = "accommodation"
    CATERING = "catering"

def _generate_excel_from_worklogs(worklogs: list[models.WorkLog], title: str) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = title

    # Headers
    headers = ["Data", "Pracownik", "Zespół", "Budowa", "Godziny", "Posiłki", "Noclegi"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)

    # Data
    for row_idx, log in enumerate(worklogs, 2):
        worker_name = log.team_member.name if log.team_member else (log.user.full_name or log.user.email)
        team_name = log.team_member.team.name if log.team_member and log.team_member.team else (log.user.team.name if log.user.team else "")
        project_code = log.project.code if log.project else ""
        
        date_val = log.date
        if isinstance(date_val, datetime):
            date_val = date_val.replace(tzinfo=None)
        ws.cell(row=row_idx, column=1, value=date_val)
        ws.cell(row=row_idx, column=2, value=worker_name)
        ws.cell(row=row_idx, column=3, value=team_name)
        ws.cell(row=row_idx, column=4, value=project_code)
        ws.cell(row=row_idx, column=5, value=log.hours_worked)
        ws.cell(row=row_idx, column=6, value=log.meals_served)
        ws.cell(row=row_idx, column=7, value=log.overnight_stays)

    # Auto-width
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

@router.get("/export")
async def export_data(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
    type: ReportType,
    project_id: int | None = None,
    team_id: int | None = None,
    company_id: int | None = None,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
):
    from openpyxl import Workbook
    stream = None
    filename = f"raport_{type.value}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    if type == ReportType.PARTICIPANTS:
        if not project_id:
            raise HTTPException(status_code=400, detail="project_id is required for participants export")
            
        # Re-use logic for participants query
        query = (
            db.query(models.WorkLog)
            .options(joinedload(models.WorkLog.team_member), joinedload(models.WorkLog.user))
            .filter(models.WorkLog.project_id == project_id)
        )
        if current_user.role != "admin":
            query = query.filter(models.WorkLog.user_id == current_user.id)
            
        worklogs = query.all()
        names = set()
        for log in worklogs:
            if log.team_member and log.team_member.name:
                names.add(log.team_member.name)
            elif log.user.full_name:
                names.add(log.user.full_name)
            else:
                names.add(log.user.email)
        sorted_names = sorted(list(names))

        wb = Workbook()
        ws = wb.active
        ws.title = "Uczestnicy"
        ws.cell(row=1, column=1, value="Nazwisko i Imię")
        for idx, name in enumerate(sorted_names, 2):
            ws.cell(row=idx, column=1, value=name)
            
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

    elif type == ReportType.TEAM_WORK:
        if not team_id:
             raise HTTPException(status_code=400, detail="team_id is required for team_work export")
        
        base_query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
        query = (
            base_query
            .join(models.WorkLog.team_member, isouter=True)
            .join(models.WorkLog.user, isouter=True)
            .filter(
                (models.TeamMember.team_id == team_id) | 
                ((models.WorkLog.team_member_id == None) & (models.User.team_id == team_id))
            )
            .options(joinedload(models.WorkLog.project))
        )
        worklogs = query.all()
        stream = _generate_excel_from_worklogs(worklogs, "Praca Zespołu")

    elif type == ReportType.ACCOMMODATION:
        query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
        query = query.filter(models.WorkLog.overnight_stays > 0).options(joinedload(models.WorkLog.project))
        if company_id:
            query = query.filter(models.WorkLog.accommodation_company_id == company_id)
        worklogs = query.all()
        stream = _generate_excel_from_worklogs(worklogs, "Noclegi")

    elif type == ReportType.CATERING:
        query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
        query = query.filter(models.WorkLog.meals_served > 0).options(joinedload(models.WorkLog.project))
        if company_id:
            query = query.filter(models.WorkLog.catering_company_id == company_id)
        worklogs = query.all()
        stream = _generate_excel_from_worklogs(worklogs, "Posiłki")

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

def calculate_easter_date(year):
    """Calculate date of Catholic Easter Sunday."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day).date()

def get_polish_holidays(year):
    """Return a set of Polish public holidays for a given year."""
    from datetime import timedelta, date
    
    easter = calculate_easter_date(year)
    easter_monday = easter + timedelta(days=1)
    corpus_christi = easter + timedelta(days=60)
    
    fixed_holidays = {
        (1, 1),   # New Year
        (1, 6),   # Epiphany
        (5, 1),   # Labor Day
        (5, 3),   # Constitution Day
        (8, 15),  # Assumption of Mary
        (11, 1),  # All Saints' Day
        (11, 11), # Independence Day
        (12, 25), # Christmas Day
        (12, 26), # Second Day of Christmas
    }
    
    holidays = {date(year, m, d) for m, d in fixed_holidays}
    holidays.add(easter)
    holidays.add(easter_monday)
    holidays.add(corpus_christi)
    
    return holidays

@router.get("/monthly-excel")
async def export_monthly_excel(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
    year: int = Query(..., description="Year of the report"),
    month: int = Query(..., description="Month of the report"),
    project_id: int | None = Query(default=None, description="Project ID filter"),
):
    # Calculate date range
    try:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid year or month")
        
    # Query Data using shared logic from worklogs.py
    base_query = _build_user_worklog_query(db, current_user, project_id, start_date, end_date)
    
    query = (
        base_query
        .options(joinedload(models.WorkLog.user).joinedload(models.User.team))
        .options(joinedload(models.WorkLog.team_member).joinedload(models.TeamMember.team))
        .options(joinedload(models.WorkLog.project))
        .options(joinedload(models.WorkLog.catering_company))
        .options(joinedload(models.WorkLog.accommodation_company))
    )
    
    worklogs = query.all()
    
    # Organize data
    user_data = {}
    
    # Daily Headers Data
    daily_headers = {d: {"projs": set(), "accs": set(), "meals": set()} for d in range(1, 32)}
    
    # Global Aggregators for Summary Sheet
    acc_global_summary = {} # Name -> count
    cat_global_summary = {} # Name -> count
    proj_global_summary = {} # Project Name -> {"meals": 0, "acc": 0, "hours": 0.0}
    
    for log in worklogs:
        if log.team_member_id:
            key = f"tm_{log.team_member_id}"
            worker_name = log.team_member.name
            team_obj = log.team_member.team
            rate = 0.0 
        else:
            key = f"u_{log.user_id}"
            worker_name = log.user.full_name or log.user.email
            team_obj = log.user.team
            rate = log.user.hourly_rate or 0.0

        if key not in user_data:
            user_data[key] = {
                "name": worker_name,
                "team": team_obj,
                "rate": rate,
                "days": {},         # day_int: hours (accumulated) or absence_note
                "project_stats": {}, # project_code -> {name, hours, meals, acc}
                "meals_count": {},  # day_int: count
                "acc_count": {},    # day_int: count
            }
        
        day = log.date.day
        
        # Populate Daily Headers
        if log.project:
            proj_str = f"{log.project.code} {log.project.name}"
            daily_headers[day]["projs"].add(proj_str)
        if log.accommodation_company:
            daily_headers[day]["accs"].add(log.accommodation_company.name)
        if log.catering_company:
            daily_headers[day]["meals"].add(log.catering_company.name)
            
        # User Data
        if log.absences > 0:
            user_data[key]["days"][day] = log.notes or "NB" # NB -> Nieobecność
        else:
            current_val = user_data[key]["days"].get(day, 0)
            if isinstance(current_val, (int, float)):
                user_data[key]["days"][day] = current_val + log.hours_worked
        
        if log.project:
            proj_code = log.project.code
            proj_name_str = log.project.name
            full_proj_name = f"{proj_code} {proj_name_str}"
            
            # Init global project summary if new
            if full_proj_name not in proj_global_summary:
                proj_global_summary[full_proj_name] = {"meals": 0, "acc": 0, "hours": 0.0}
            
            proj_global_summary[full_proj_name]["hours"] += log.hours_worked

            # Init user project stats if new
            if proj_code not in user_data[key]["project_stats"]:
                user_data[key]["project_stats"][proj_code] = {
                    "name": proj_name_str,
                    "hours": 0,
                    "meals": 0,
                    "acc": 0
                }
            
            # Accumulate User Project Stats
            user_data[key]["project_stats"][proj_code]["hours"] += log.hours_worked
            if log.meals_served > 0:
                 user_data[key]["project_stats"][proj_code]["meals"] += log.meals_served
            if log.overnight_stays > 0:
                 user_data[key]["project_stats"][proj_code]["acc"] += log.overnight_stays

        # Meals
        if log.meals_served > 0:
            user_data[key]["meals_count"][day] = user_data[key]["meals_count"].get(day, 0) + log.meals_served
            
            name = log.catering_company.name if log.catering_company else "Tak"
            # Global Aggregation
            cat_global_summary[name] = cat_global_summary.get(name, 0) + log.meals_served
            if log.project:
                full_proj_name = f"{log.project.code} {log.project.name}"
                if full_proj_name in proj_global_summary:
                    proj_global_summary[full_proj_name]["meals"] += log.meals_served

        # Accommodation
        if log.overnight_stays > 0:
            user_data[key]["acc_count"][day] = user_data[key]["acc_count"].get(day, 0) + log.overnight_stays
            
            name = log.accommodation_company.name if log.accommodation_company else "Tak"
            # Global Aggregation
            acc_global_summary[name] = acc_global_summary.get(name, 0) + log.overnight_stays
            if log.project:
                full_proj_name = f"{log.project.code} {log.project.name}"
                if full_proj_name in proj_global_summary:
                    proj_global_summary[full_proj_name]["acc"] += log.overnight_stays

    # Create Workbook In-Memory
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Lista"

    # Set Header (Month/Year) - Cell B1
    polish_months = {
        1: "STYCZEŃ", 2: "LUTY", 3: "MARZEC", 4: "KWIECIEŃ", 5: "MAJ", 6: "CZERWIEC",
        7: "LIPIEC", 8: "SIERPIEŃ", 9: "WRZESIEŃ", 10: "PAŹDZIERNIK", 11: "LISTOPAD", 12: "GRUDZIEŃ"
    }
    month_name = polish_months.get(month, "").upper()
    ws["B1"] = f"{month_name} {year}"
    ws["B1"].font = Font(bold=True, size=14)

    # Styles
    thin_side = Side(style='thin')
    medium_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Fills
    fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    holidays = get_polish_holidays(year)
    day_styles = {}
    
    # Pre-calculate day styles and count working days
    total_working_days = 0
    for day in range(1, 32):
        try:
            current_date = datetime(year, month, day).date()
            if current_date in holidays or current_date.weekday() == 6: # Sunday
                day_styles[day] = fill_red
            elif current_date.weekday() == 5: # Saturday
                day_styles[day] = fill_yellow
            else:
                day_styles[day] = None
                total_working_days += 1
        except ValueError:
             day_styles[day] = None

    # Columns: 
    # A: Lp, B: Name, C: Days, D: Rate, E: Type (Kod/G/N/P)
    # F starts Days (day 1)
    
    # --- HEADERS ---
    
    # Header Labels (Col A-D)
    # Row 2: Acc Firm
    # Row 3: Meal Firm
    # Row 4: Day Num
    
    ws.cell(row=2, column=2, value="Firma nocleg").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Firma posiłek").font = Font(bold=True)
    
    # Fill Daily Info (Header)
    for day in range(1, 32):
        col_idx = 5 + day # F is 6
        
        # Row 2: Accs
        accs = daily_headers[day]["accs"]
        ws.cell(row=2, column=col_idx, value="\n".join(accs)).alignment = center_align
        ws.cell(row=2, column=col_idx).border = medium_border

        # Row 3: Meals
        meals = daily_headers[day]["meals"]
        ws.cell(row=3, column=col_idx, value="\n".join(meals)).alignment = center_align
        ws.cell(row=3, column=col_idx).border = medium_border
        
        # Row 4: Day Num
        c = ws.cell(row=4, column=col_idx, value=day)
        c.border = medium_border
        c.alignment = center_align
        c.font = Font(bold=True)
        if day_styles.get(day): c.fill = day_styles[day]

    # Fixed Header Labels (Row 4 - Main Table Header)
    # Lp
    c = ws.cell(row=4, column=1, value="Lp")
    c.border = medium_border
    c.alignment = center_align
    # Name
    c = ws.cell(row=4, column=2, value="Nazwisko i Imię")
    c.border = medium_border
    c.alignment = center_align
    # Days
    c = ws.cell(row=4, column=3, value="Liczba dni")
    c.border = medium_border
    c.alignment = center_align
    # Rate
    c = ws.cell(row=4, column=4, value="Stawka")
    c.border = medium_border
    c.alignment = center_align
    # Type
    c = ws.cell(row=4, column=5, value="Typ")
    c.border = medium_border
    c.alignment = center_align

    # Summary Headers (after day 31 -> Col 37/AK)
    summary_start_col = 6 + 31 # 37
    
    summary_titles = ["ILOŚĆ GODZIN", "Noclegi", "Posiłki", "NALEŻNOŚĆ"]
    for i, title in enumerate(summary_titles):
        c = ws.cell(row=4, column=summary_start_col + i, value=title)
        c.border = medium_border
        c.alignment = center_align
        
    payroll_headers = [
        "premia (DODATEK)", "PREMIA PRZELEW", "PŁACA NA KONTO", 
        "Komornik", "PPK", "ubezp. Pak. Med.", 
        "PREMIA LISTA", "PŁACA GOTÓWKA", "PREMIA GOTÓWKA", 
        "WYPŁATY POKRYCIA", "POTRĄCENIA", "DO WYPŁATY"
    ]
    for i, title in enumerate(payroll_headers):
        c = ws.cell(row=4, column=summary_start_col + 4 + i, value=title)
        c.border = medium_border
        c.alignment = center_align
        
    # --- USER DATA ---
    users_sorted = sorted(user_data.values(), key=lambda x: (
        (x["team"].name if x["team"] else ""), 
        (x["name"] or "")
    ))
    
    current_row = 5
    for idx, item in enumerate(users_sorted):
        is_gray = (idx % 2 == 1)
        base_fill = fill_gray if is_gray else None
        
        start_r = current_row
        end_r = current_row + 3 # 4 rows: Kod, G, N, P
        
        # Merge Fixed Columns (A-D)
        # Lp (A)
        ws.merge_cells(start_row=start_r, start_column=1, end_row=end_r, end_column=1)
        c = ws.cell(row=start_r, column=1, value=idx + 1)
        c.border = medium_border
        c.alignment = center_align
        if base_fill: c.fill = base_fill
        
        # Name (B)
        ws.merge_cells(start_row=start_r, start_column=2, end_row=end_r, end_column=2)
        c = ws.cell(row=start_r, column=2, value=item["name"])
        c.border = medium_border
        c.alignment = left_align
        if base_fill: c.fill = base_fill
        
        # Days Work (C)
        ws.merge_cells(start_row=start_r, start_column=3, end_row=end_r, end_column=3)
        c = ws.cell(row=start_r, column=3, value=total_working_days)
        c.border = medium_border
        c.alignment = center_align
        if base_fill: c.fill = base_fill
        
        # Rate (D)
        ws.merge_cells(start_row=start_r, start_column=4, end_row=end_r, end_column=4)
        c = ws.cell(row=start_r, column=4, value=item["rate"])
        c.border = medium_border
        c.alignment = center_align
        c.number_format = '0.00'
        if base_fill: c.fill = base_fill
        
        # Type (E) - Not merged
        types = ["Budowa", "G", "N", "P"] 
        for i, t in enumerate(types):
            c = ws.cell(row=start_r + i, column=5, value=t)
            c.border = medium_border
            c.alignment = center_align
            if base_fill: c.fill = base_fill
            
        # Daily Data (F...)
        for day in range(1, 32):
            col_idx = 5 + day
            d_fill = day_styles.get(day) or base_fill
            
            # Row 1: Budowa (Project)
            proj_str = item["projects"].get(day, "")
            c = ws.cell(row=start_r, column=col_idx, value=proj_str)
            c.border = medium_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if d_fill: c.fill = d_fill
            
            # Row 2: G (Hours)
            val_h = item["days"].get(day, "")
            c = ws.cell(row=start_r + 1, column=col_idx, value=val_h)
            c.border = medium_border
            c.alignment = center_align
            if d_fill: c.fill = d_fill
            
            # Row 3: N - Acc (Tak/Nie)
            count_n = item["acc_count"].get(day, 0)
            val_n = "Tak" if count_n > 0 else ""
            c = ws.cell(row=start_r + 2, column=col_idx, value=val_n)
            c.border = medium_border
            c.alignment = center_align
            if d_fill: c.fill = d_fill
            
            # Row 4: P - Meals (Tak/Nie)
            count_p = item["meals_count"].get(day, 0)
            val_p = "Tak" if count_p > 0 else ""
            c = ws.cell(row=start_r + 3, column=col_idx, value=val_p)
            c.border = medium_border
            c.alignment = center_align
            if d_fill: c.fill = d_fill
            
        # Summary Formulas
        # Row 1 (Budowa) - No summary
        
        # Row 2 (G)
        row_g = start_r + 1
        # Row 3 (N) 
        row_n = start_r + 2
        # Row 4 (P)
        row_p = start_r + 3
        
        last_day_col_let = get_column_letter(6 + 30) # AJ (F is 6. 6+30=36 -> AJ)
        range_g = f"F{row_g}:{last_day_col_let}{row_g}"
        range_n = f"F{row_n}:{last_day_col_let}{row_n}"
        range_p = f"F{row_p}:{last_day_col_let}{row_p}"
        
        # Hours
        form_h = f"=SUM({range_g})"
        c = ws.cell(row=start_r, column=summary_start_col, value=form_h) # Top-align summary with user block?
        # Actually summary should probably be merged across 4 rows as well
        
        # Note: Merging Formula cells can be tricky if not careful, but usually works if value is in top-left.
        # I'll put values in the cells corresponding to the rows, or just merge all 4.
        
        for k in range(4):
            col = summary_start_col + k
            ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
            
            c = ws.cell(row=start_r, column=col)
            c.border = medium_border
            c.alignment = center_align
            if base_fill: c.fill = base_fill
            
            # Assign formula to the merged text cell
            if k == 0: # Hours
                c.value = form_h
            elif k == 1: # Acc
                 c.value = f'=COUNTIF({range_n}, "Tak")'
            elif k == 2: # Meals
                 c.value = f'=COUNTIF({range_p}, "Tak")'
            elif k == 3: # Pay
                 # Pay = total_hours * rate
                 # Rate is D{start_r} (merged) -> D{start_r} is safe
                 # Total Hours is at this column (summary_start_col){start_r}
                 c_hours_cell = f"{get_column_letter(summary_start_col)}{start_r}"
                 c.value = f"={c_hours_cell}*D{start_r}"
                 c.number_format = '0.00'

        # Payroll Columns (Empty) - Merged
        for k in range(len(payroll_headers)):
            col = summary_start_col + 4 + k
            ws.merge_cells(start_row=start_r, start_column=col, end_row=end_r, end_column=col)
            for r in range(start_r, end_r + 1):
                c = ws.cell(row=r, column=col)
                c.border = medium_border
                if base_fill: c.fill = base_fill

        current_row += 4
        
    # Auto-fit (Roughly)
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 5
    for day in range(1, 32):
        ws.column_dimensions[get_column_letter(5 + day)].width = 5

    # --- SUMMARY SHEET ("Podsumowanie") ---
    ws_summary = wb.create_sheet("Podsumowanie")
    
    # helper for borders
    def set_border(cell):
        cell.border = medium_border
        cell.alignment = center_align

    # 1. Catering Summary
    row_idx = 2
    ws_summary.cell(row=row_idx, column=2, value="PODSUMOWANIE POSIŁKÓW").font = Font(bold=True)
    row_idx += 1
    
    ws_summary.cell(row=row_idx, column=2, value="Firma")
    set_border(ws_summary.cell(row=row_idx, column=2))
    ws_summary.cell(row=row_idx, column=3, value="Ilość")
    set_border(ws_summary.cell(row=row_idx, column=3))
    row_idx += 1
    
    total_meals = 0
    for company, count in cat_global_summary.items():
        ws_summary.cell(row=row_idx, column=2, value=company or "Brak firmy")
        set_border(ws_summary.cell(row=row_idx, column=2))
        
        ws_summary.cell(row=row_idx, column=3, value=count)
        set_border(ws_summary.cell(row=row_idx, column=3))
        total_meals += count
        row_idx += 1
        
    # Total row for catering
    ws_summary.cell(row=row_idx, column=2, value="RAZEM").font = Font(bold=True)
    set_border(ws_summary.cell(row=row_idx, column=2))
    ws_summary.cell(row=row_idx, column=3, value=total_meals).font = Font(bold=True)
    set_border(ws_summary.cell(row=row_idx, column=3))
    
    row_idx += 3 # Gap
    
    # 2. Accommodation Summary
    ws_summary.cell(row=row_idx, column=2, value="PODSUMOWANIE NOCLEGÓW").font = Font(bold=True)
    row_idx += 1
    
    ws_summary.cell(row=row_idx, column=2, value="Firma")
    set_border(ws_summary.cell(row=row_idx, column=2))
    ws_summary.cell(row=row_idx, column=3, value="Ilość")
    set_border(ws_summary.cell(row=row_idx, column=3))
    row_idx += 1
    
    total_acc = 0
    for company, count in acc_global_summary.items():
        ws_summary.cell(row=row_idx, column=2, value=company or "Brak firmy")
        set_border(ws_summary.cell(row=row_idx, column=2))
        
        ws_summary.cell(row=row_idx, column=3, value=count)
        set_border(ws_summary.cell(row=row_idx, column=3))
        total_acc += count
        row_idx += 1

    # Total row for accommodation
    ws_summary.cell(row=row_idx, column=2, value="RAZEM").font = Font(bold=True)
    set_border(ws_summary.cell(row=row_idx, column=2))
    ws_summary.cell(row=row_idx, column=3, value=total_acc).font = Font(bold=True)
    set_border(ws_summary.cell(row=row_idx, column=3))

    row_idx += 3 # Gap

    # 3. Project Summary
    ws_summary.cell(row=row_idx, column=2, value="PODSUMOWANIE PROJEKTÓW").font = Font(bold=True)
    row_idx += 1
    
    headers = ["Projekt", "Godziny", "% Godzin", "Posiłki", "% Posiłków", "Noclegi", "% Noclegów"]
    for i, h in enumerate(headers):
        c = ws_summary.cell(row=row_idx, column=2+i, value=h)
        set_border(c)
        c.font = Font(bold=True)
    row_idx += 1
    
    total_hours_global = sum(d["hours"] for d in proj_global_summary.values())

    for proj_name, data in proj_global_summary.items():
        c = ws_summary.cell(row=row_idx, column=2, value=proj_name)
        set_border(c)
        
        hours_sum = data["hours"]
        c = ws_summary.cell(row=row_idx, column=3, value=hours_sum)
        set_border(c)

        pct_hours = (hours_sum / total_hours_global) if total_hours_global > 0 else 0
        c = ws_summary.cell(row=row_idx, column=4, value=pct_hours)
        c.number_format = '0.00%'
        set_border(c)

        meals_count = data["meals"]
        c = ws_summary.cell(row=row_idx, column=5, value=meals_count)
        set_border(c)
        
        pct_meals = (meals_count / total_meals) if total_meals > 0 else 0
        c = ws_summary.cell(row=row_idx, column=6, value=pct_meals)
        c.number_format = '0.00%'
        set_border(c)
        
        acc_count = data["acc"]
        c = ws_summary.cell(row=row_idx, column=7, value=acc_count)
        set_border(c)
        
        pct_acc = (acc_count / total_acc) if total_acc > 0 else 0
        c = ws_summary.cell(row=row_idx, column=8, value=pct_acc)
        c.number_format = '0.00%'
        set_border(c)
        
        row_idx += 1

    for column_cells in ws_summary.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws_summary.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # --- EMPLOYEE SUMMARY SHEET ("Podsumowanie pracowników") ---
    ws_emp = wb.create_sheet("Podsumowanie pracowników")
    
    row_idx = 1
    all_projects = sorted(proj_global_summary.keys())
    
    for item in users_sorted:
        days_worked = sum(1 for v in item["days"].values() if isinstance(v, (int, float)) and v > 0)
        absences = sum(1 for v in item["days"].values() if isinstance(v, str))
        total_user_hours = sum(p["hours"] for p in item["project_stats"].values())
        
        c = ws_emp.cell(row=row_idx, column=1, value=item['name'])
        c.font = Font(bold=True)
        set_border(c)
        row_idx += 1
        
        ws_emp.cell(row=row_idx, column=1, value="Kategoria")
        set_border(ws_emp.cell(row=row_idx, column=1))
        
        ws_emp.cell(row=row_idx, column=2, value="Dni pracy")
        set_border(ws_emp.cell(row=row_idx, column=2))
        ws_emp.cell(row=row_idx, column=3, value="Nieobecności")
        set_border(ws_emp.cell(row=row_idx, column=3))
        ws_emp.cell(row=row_idx, column=4, value="Suma godzin")
        set_border(ws_emp.cell(row=row_idx, column=4))
        ws_emp.cell(row=row_idx, column=5, value="% Godzin")
        set_border(ws_emp.cell(row=row_idx, column=5))
        
        for i, proj in enumerate(all_projects):
            c = ws_emp.cell(row=row_idx, column=6+i, value=proj)
            set_border(c)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row_idx += 1
        
        c = ws_emp.cell(row=row_idx, column=1, value="Godziny")
        set_border(c)
        
        c = ws_emp.cell(row=row_idx, column=2, value=f"{days_worked}/{total_working_days}") 
        set_border(c)
        c = ws_emp.cell(row=row_idx, column=3, value=absences) 
        set_border(c)
        c = ws_emp.cell(row=row_idx, column=4, value=total_user_hours) 
        set_border(c)
        
        user_share = (total_user_hours / total_hours_global) if total_hours_global > 0 else 0
        c = ws_emp.cell(row=row_idx, column=5, value=user_share)
        c.number_format = '0.00%'
        set_border(c)

        for i, proj in enumerate(all_projects):
            p_code_extracted = proj.split(" ")[0]
            val = item["project_stats"].get(p_code_extracted, {}).get("hours", 0)
            c = ws_emp.cell(row=row_idx, column=6+i, value=val)
            set_border(c)
        row_idx += 1
        
        c = ws_emp.cell(row=row_idx, column=1, value="% Czasu")
        set_border(c)
        
        for k in range(2, 6): set_border(ws_emp.cell(row=row_idx, column=k))
        
        for i, proj in enumerate(all_projects):
            p_code_extracted = proj.split(" ")[0]
            val = item["project_stats"].get(p_code_extracted, {}).get("hours", 0)
            pct = (val / total_user_hours) if total_user_hours > 0 else 0
            
            c = ws_emp.cell(row=row_idx, column=6+i, value=pct)
            c.number_format = '0.00%'
            set_border(c)
        row_idx += 1
        
        c = ws_emp.cell(row=row_idx, column=1, value="Posiłki")
        set_border(c)
        
        for k in range(2, 6): set_border(ws_emp.cell(row=row_idx, column=k))
        
        for i, proj in enumerate(all_projects):
            p_code_extracted = proj.split(" ")[0]
            val = item["project_stats"].get(p_code_extracted, {}).get("meals", 0)
            c = ws_emp.cell(row=row_idx, column=6+i, value=val)
            set_border(c)
        row_idx += 1
            
        c = ws_emp.cell(row=row_idx, column=1, value="Noclegi")
        set_border(c)

        for k in range(2, 6): set_border(ws_emp.cell(row=row_idx, column=k))

        for i, proj in enumerate(all_projects):
            p_code_extracted = proj.split(" ")[0]
            val = item["project_stats"].get(p_code_extracted, {}).get("acc", 0)
            c = ws_emp.cell(row=row_idx, column=6+i, value=val)
            set_border(c)
        row_idx += 1
        
        row_idx += 1

    ws_emp.column_dimensions['A'].width = 30 
    ws_emp.column_dimensions['B'].width = 12
    ws_emp.column_dimensions['C'].width = 15
    ws_emp.column_dimensions['D'].width = 12
    ws_emp.column_dimensions['E'].width = 10
    for i, _ in enumerate(all_projects):
         ws_emp.column_dimensions[get_column_letter(6+i)].width = 15 

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"raport_{year}_{month:02d}.xlsx"
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/participants", response_model=list[str])
async def get_project_participants(
    project_id: int,
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
):
    # Query distinct names from worklogs for a project
    # We prioritize TeamMember names, fallback to User names
    # Effectively we want unique names of people who worked on the project
    
    # Using specific query to get distinct names
    query = (
        db.query(models.WorkLog)
        .options(joinedload(models.WorkLog.team_member), joinedload(models.WorkLog.user))
        .filter(models.WorkLog.project_id == project_id)
    )
    
    if current_user.role != "admin":
        # Non-admins restricted to their own logs
        query = query.filter(models.WorkLog.user_id == current_user.id)
        
    worklogs = query.all()
    
    names = set()
    for log in worklogs:
        if log.team_member and log.team_member.name:
            names.add(log.team_member.name)
        elif log.user.full_name:
            names.add(log.user.full_name)
        else:
            names.add(log.user.email)
            
    return sorted(list(names))

@router.get("/team-work", response_model=schemas.PaginatedResponse[schemas.WorkLogRead])
async def get_team_work_report(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
    team_id: int,
    start_date: datetime | None = Query(default=None),
    end_date: datetime | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=100),
):
    # Retrieve logs for a specific team. 
    # Logic: Log associated with a TeamMember of that team OR User of that team
    base_query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
    
    # Join necessary tables to filter by team
    query = (
        base_query
        .join(models.WorkLog.team_member, isouter=True)
        .join(models.WorkLog.user, isouter=True)
        .filter(
            (models.TeamMember.team_id == team_id) | 
            ((models.WorkLog.team_member_id == None) & (models.User.team_id == team_id))
        )
    )
    
    total = query.count()
    pages = (total + size - 1) // size
    items = query.offset((page - 1) * size).limit(size).all()
    
    return schemas.PaginatedResponse(items=items, total=total, page=page, size=size, pages=pages)

@router.get("/accommodation", response_model=schemas.PaginatedResponse[schemas.WorkLogRead])
async def get_accommodation_report(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
    company_id: int | None = Query(default=None),
    start_date: datetime | None = Query(default=None),
    end_date: datetime | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=100),
):
    query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
    
    # Filter for logs with accommodation
    query = query.filter(models.WorkLog.overnight_stays > 0)
    
    if company_id:
        query = query.filter(models.WorkLog.accommodation_company_id == company_id)
        
    total = query.count()
    pages = (total + size - 1) // size
    items = query.offset((page - 1) * size).limit(size).all()
    
    return schemas.PaginatedResponse(items=items, total=total, page=page, size=size, pages=pages)

@router.get("/catering", response_model=schemas.PaginatedResponse[schemas.WorkLogRead])
async def get_catering_report(
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[models.User, Depends(auth.get_current_active_user)],
    company_id: int | None = Query(default=None),
    start_date: datetime | None = Query(default=None),
    end_date: datetime | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    size: int = Query(default=50, ge=1, le=100),
):
    query = _build_user_worklog_query(db, current_user, None, start_date, end_date)
    
    # Filter for logs with meals
    query = query.filter(models.WorkLog.meals_served > 0)
    
    if company_id:
        query = query.filter(models.WorkLog.catering_company_id == company_id)
        
    total = query.count()
    pages = (total + size - 1) // size
    items = query.offset((page - 1) * size).limit(size).all()
    
    return schemas.PaginatedResponse(items=items, total=total, page=page, size=size, pages=pages)
